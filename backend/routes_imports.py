"""Excel/CSV candidate import: preview (auto column mapping) -> commit (migrate rows into ATS)."""
import csv
import io
import re
from datetime import datetime, timezone
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from fastapi.responses import Response
from pydantic import BaseModel

from auth import require_roles
from database import db
from utils import clean, log_activity, log_audit, new_id, next_candidate_code, now_iso

router = APIRouter(prefix='/imports', tags=['imports'])

MAX_ROWS = 5000

TARGET_FIELDS = [
    'name', 'email', 'phone', 'current_title', 'current_company', 'location',
    'skills', 'job', 'stage', 'source', 'tags', 'applied_at', 'notice_period', 'notes',
]

# header synonyms for auto-mapping (checked in order: exact match first, then contains)
SYNONYMS = {
    'name': ['name', 'full name', 'candidate name', 'candidate', 'applicant name', 'applicant'],
    'email': ['email', 'e-mail', 'email address', 'mail', 'email id'],
    'phone': ['phone', 'phone number', 'mobile', 'mobile number', 'contact number', 'telephone', 'contact'],
    'current_title': ['current title', 'job title', 'title', 'position', 'designation', 'current role', 'role'],
    'current_company': ['current company', 'company', 'employer', 'organization', 'organisation', 'current employer'],
    'location': ['location', 'city', 'address', 'place', 'current location'],
    'skills': ['skills', 'skill set', 'skillset', 'technologies', 'tech stack', 'key skills'],
    'job': ['job applied', 'applied for', 'applying for', 'position applied', 'job opening', 'requisition', 'vacancy', 'opening', 'job'],
    'stage': ['stage', 'pipeline stage', 'current stage', 'status', 'hiring stage'],
    'source': ['source', 'channel', 'referral source', 'sourced from', 'application source'],
    'tags': ['tags', 'labels', 'tag'],
    'applied_at': ['applied date', 'application date', 'date applied', 'applied on', 'applied at', 'date of application', 'date'],
    'notice_period': ['notice period', 'notice', 'availability', 'available from', 'availability to join'],
    'notes': ['notes', 'comments', 'remarks', 'note', 'comment'],
}

EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')

SOURCE_MAP = {
    'referral': 'referral', 'employee referral': 'referral', 'refer': 'referral',
    'job board': 'job_board', 'job_board': 'job_board', 'jobboard': 'job_board', 'indeed': 'job_board', 'naukri': 'job_board', 'monster': 'job_board',
    'career site': 'career_site', 'career_site': 'career_site', 'careers page': 'career_site', 'website': 'career_site', 'career page': 'career_site',
    'linkedin': 'linkedin', 'linked in': 'linkedin',
}

DATE_FORMATS = ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y', '%d %b %Y', '%b %d, %Y', '%d.%m.%Y', '%Y/%m/%d']


def _cell_str(v) -> str:
    if v is None:
        return ''
    if isinstance(v, datetime):
        return v.strftime('%Y-%m-%d')
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def _parse_file(data: bytes, filename: str):
    ext = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if ext == 'xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([_cell_str(c) for c in row])
            if len(rows) > MAX_ROWS + 1:
                break
        wb.close()
    elif ext == 'csv':
        text = data.decode('utf-8-sig', errors='ignore')
        reader = csv.reader(io.StringIO(text))
        rows = []
        for row in reader:
            rows.append([_cell_str(c) for c in row])
            if len(rows) > MAX_ROWS + 1:
                break
    else:
        raise ValueError('Unsupported file type. Please upload .xlsx or .csv')
    # drop fully-empty rows
    rows = [r for r in rows if any(c for c in r)]
    if len(rows) < 2:
        raise ValueError('File must contain a header row and at least one data row')
    headers = [h if h else f'Column {i + 1}' for i, h in enumerate(rows[0])]
    data_rows = rows[1:]
    # normalize row lengths
    width = len(headers)
    data_rows = [(r + [''] * width)[:width] for r in data_rows]
    return headers, data_rows


def _suggest_mapping(headers):
    mapping = {}
    used = set()
    norm = [h.strip().lower() for h in headers]
    # pass 1: exact synonym match
    for field, syns in SYNONYMS.items():
        for i, h in enumerate(norm):
            if headers[i] in mapping or field in used:
                continue
            if h in syns:
                mapping[headers[i]] = field
                used.add(field)
                break
    # pass 2: contains match
    for field, syns in SYNONYMS.items():
        if field in used:
            continue
        for i, h in enumerate(norm):
            if headers[i] in mapping:
                continue
            if any(s in h for s in syns):
                mapping[headers[i]] = field
                used.add(field)
                break
    for h in headers:
        mapping.setdefault(h, 'skip')
    return mapping


def _split_list(val: str):
    return [s.strip() for s in re.split(r'[,;|/]', val or '') if s.strip()]


def _parse_date(val: str) -> Optional[str]:
    if not val:
        return None
    v = val.strip()
    try:
        return datetime.fromisoformat(v.replace('Z', '+00:00')).astimezone(timezone.utc).isoformat()
    except ValueError:
        pass
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(v, fmt).replace(tzinfo=timezone.utc).isoformat()
        except ValueError:
            continue
    return None


@router.get('/template')
async def download_template(user: dict = Depends(require_roles('admin', 'recruiter'))):
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Candidates'
    headers = ['Name', 'Email', 'Phone', 'Current Title', 'Current Company', 'Location',
               'Skills', 'Job Applied', 'Stage', 'Source', 'Tags', 'Applied Date', 'Notice Period', 'Notes']
    ws.append(headers)
    ws.append(['Jane Doe', 'jane.doe@example.com', '(555) 123-4567', 'Software Engineer', 'Acme Inc.',
               'San Francisco, CA', 'Python; React; SQL', 'Senior Backend Engineer', 'Applied',
               'LinkedIn', 'remote-ok; senior', '2026-01-15', '30 days', 'Strong referral from team lead'])
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = max(14, len(h) + 4)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(
        content=buf.getvalue(),
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': 'attachment; filename="candidate_import_template.xlsx"'},
    )


@router.post('/preview')
async def preview_import(file: UploadFile = File(...), user: dict = Depends(require_roles('admin', 'recruiter'))):
    data = await file.read()
    if len(data) > 15 * 1024 * 1024:
        raise HTTPException(status_code=422, detail='File exceeds 15MB limit')
    try:
        headers, rows = _parse_file(data, file.filename)
    except ValueError as e:
        raise HTTPException(status_code=422, detail=str(e))
    except Exception:
        raise HTTPException(status_code=422, detail='Could not read this file. Ensure it is a valid .xlsx or .csv')
    if len(rows) > MAX_ROWS:
        raise HTTPException(status_code=422, detail=f'File has too many rows (max {MAX_ROWS})')
    session = {
        'id': new_id(),
        'filename': file.filename,
        'headers': headers,
        'rows': rows,
        'total_rows': len(rows),
        'created_by': user['id'],
        'created_at': now_iso(),
        'status': 'pending',
    }
    await db.import_sessions.insert_one(session)
    return {
        'import_id': session['id'],
        'filename': file.filename,
        'headers': headers,
        'suggested_mapping': _suggest_mapping(headers),
        'sample_rows': rows[:5],
        'total_rows': len(rows),
        'target_fields': TARGET_FIELDS,
    }


class CommitBody(BaseModel):
    mapping: dict  # header -> target field or 'skip'
    default_job_id: Optional[str] = None
    default_source: str = 'career_site'
    duplicate_strategy: str = 'skip'  # skip | create


@router.post('/{import_id}/commit')
async def commit_import(import_id: str, body: CommitBody, user: dict = Depends(require_roles('admin', 'recruiter'))):
    session = await db.import_sessions.find_one({'id': import_id})
    if not session:
        raise HTTPException(status_code=404, detail='Import session not found or expired')
    if session.get('status') == 'committed':
        raise HTTPException(status_code=409, detail='This import has already been committed')

    headers = session['headers']
    rows = session['rows']
    # header index -> target field
    col_map = {}
    for i, h in enumerate(headers):
        target = body.mapping.get(h, 'skip')
        if target in TARGET_FIELDS:
            col_map[i] = target
    if 'name' not in col_map.values():
        raise HTTPException(status_code=422, detail='A column must be mapped to Name')

    jobs = await db.jobs.find({}, {'_id': 0}).to_list(500)
    jobs_by_title = {j['title'].strip().lower(): j for j in jobs}
    jobs_by_id = {j['id']: j for j in jobs}
    default_job = jobs_by_id.get(body.default_job_id) if body.default_job_id else None

    settings = await db.settings.find_one({'key': 'pipeline_stages'})
    global_stages = [s['name'] for s in settings['stages']] if settings else ['Applied', 'Screening', 'Interview', 'Offer', 'Hired', 'Rejected']

    existing_emails = set(e.lower() for e in await db.candidates.distinct('email') if e)
    seen_emails = set()

    created_docs = []
    errors = []
    skipped_duplicates = 0

    for idx, row in enumerate(rows):
        rownum = idx + 2  # 1-based + header row
        vals = {}
        for i, field in col_map.items():
            vals[field] = row[i].strip() if i < len(row) else ''

        name = vals.get('name', '')
        if not name:
            errors.append({'row': rownum, 'reason': 'Missing name — row skipped'})
            continue

        email = (vals.get('email') or '').lower() or None
        warnings = []
        if email and not EMAIL_RE.match(email):
            warnings.append('invalid email format — email left blank')
            email = None
        if email:
            if email in existing_emails or email in seen_emails:
                if body.duplicate_strategy == 'skip':
                    skipped_duplicates += 1
                    errors.append({'row': rownum, 'reason': f'Duplicate email {email} — skipped'})
                    continue
            seen_emails.add(email)

        # job matching
        job = None
        job_val = vals.get('job', '')
        if job_val:
            job = jobs_by_title.get(job_val.strip().lower())
            if not job:
                for title, j in jobs_by_title.items():
                    if job_val.strip().lower() in title or title in job_val.strip().lower():
                        job = j
                        break
        if not job:
            job = default_job
        if not job:
            errors.append({'row': rownum, 'reason': f'Could not match job "{job_val}" and no default job selected — row skipped'})
            continue

        job_stages = job.get('stages') or global_stages
        stage_val = (vals.get('stage') or '').strip()
        stage = next((s for s in job_stages if s.lower() == stage_val.lower()), None) if stage_val else None
        if stage_val and not stage:
            warnings.append(f'unknown stage "{stage_val}" — placed in {job_stages[0]}')
        if not stage:
            stage = job_stages[0]

        source_val = (vals.get('source') or '').strip().lower()
        source = SOURCE_MAP.get(source_val) or (source_val.replace(' ', '_') if source_val else body.default_source)

        applied_at = _parse_date(vals.get('applied_at', '')) or now_iso()

        status = 'active'
        hired_at = None
        rejection_reason = None
        if stage.lower() == 'hired':
            status, hired_at = 'hired', now_iso()
        elif stage.lower() == 'rejected':
            status, rejection_reason = 'rejected', 'Imported as rejected'

        low_conf = []
        if warnings:
            low_conf = [w.split(' ')[1] if w.startswith('invalid email') else 'stage' for w in warnings]

        doc = {
            'id': new_id(),
            'candidate_code': await next_candidate_code(),
            'name': name,
            'email': email,
            'phone': vals.get('phone') or None,
            'current_title': vals.get('current_title') or None,
            'current_company': vals.get('current_company') or None,
            'location': vals.get('location') or None,
            'experience': [],
            'education': [],
            'skills': _split_list(vals.get('skills', '')),
            'job_id': job['id'],
            'stage': stage,
            'source': source,
            'recruiter_id': job.get('recruiter_id') or user['id'],
            'tags': _split_list(vals.get('tags', '')),
            'resume_file_id': None,
            'low_confidence_fields': [],
            'notice_period': vals.get('notice_period') or None,
            'status': status,
            'rejection_reason': rejection_reason,
            'applied_at': applied_at,
            'hired_at': hired_at,
            'created_at': now_iso(),
            'updated_at': now_iso(),
            'imported': True,
            'import_id': import_id,
        }
        created_docs.append(doc)
        note_text = vals.get('notes')
        if note_text:
            await db.notes.insert_one({
                'id': new_id(), 'candidate_id': doc['id'], 'author_id': user['id'], 'author_name': user['name'],
                'text': f'[Imported] {note_text}', 'note_type': 'note', 'created_at': now_iso(),
            })
        if warnings:
            errors.append({'row': rownum, 'reason': f'{name}: imported with warnings — ' + '; '.join(warnings)})

    if created_docs:
        await db.candidates.insert_many(created_docs)
    await db.import_sessions.update_one({'id': import_id}, {'$set': {'status': 'committed', 'committed_at': now_iso(),
                                                                     'created_count': len(created_docs)}, '$unset': {'rows': ''}})
    await log_activity(user, 'import', f"imported {len(created_docs)} candidates from {session['filename']}")
    await log_audit(user, 'candidates_imported', 'candidate', 'bulk',
                    f"{session['filename']}: {len(created_docs)} created, {skipped_duplicates} duplicates skipped, {len(errors)} issues")

    return {
        'created': len(created_docs),
        'skipped_duplicates': skipped_duplicates,
        'errors': errors,
        'total_rows': session['total_rows'],
        'candidates': [{'id': d['id'], 'name': d['name'], 'email': d['email'], 'stage': d['stage'], 'job_id': d['job_id']} for d in created_docs[:50]],
    }
